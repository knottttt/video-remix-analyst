from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook

from common import ensure_dir, load_project, ms_to_timestamp, save_project, skill_root, template_env


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--project-dir", required=True)
    parser.add_argument("--mode", choices=["analysis"], default="analysis")
    return parser.parse_args()


def write_analysis_excel(project_dir: Path, project: dict) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Shots"
    ws.append(["shot_id", "time_start", "time_end", "beat_id", "shot_size", "camera_motion", "main_subject", "story_function", "analysis_mode", "story_description", "change_description", "emotion_progression", "transition_bridge"])
    for shot in project["shots"]:
        ws.append(
            [
                shot["shot_id"],
                ms_to_timestamp(shot["start_ms"]),
                ms_to_timestamp(shot["end_ms"]),
                shot.get("beat_id"),
                shot.get("shot_size"),
                shot.get("camera_motion"),
                shot.get("main_subject"),
                shot.get("story_function"),
                shot.get("analysis_mode"),
                shot.get("story_description"),
                shot.get("change_description"),
                shot.get("emotion_progression"),
                shot.get("transition_bridge"),
            ]
        )
    beat_ws = wb.create_sheet("Beats")
    beat_ws.append(["beat_id", "beat_type", "time_start", "time_end", "objective", "emotion", "pacing", "audio_beat_aligned", "what_happens", "story_purpose", "transition_to_next"])
    for beat in project["beats"]:
        beat_ws.append(
            [
                beat["beat_id"],
                beat["beat_type"],
                ms_to_timestamp(beat["start_ms"]),
                ms_to_timestamp(beat["end_ms"]),
                beat["objective"],
                beat["emotional_direction"],
                beat["pacing_note"],
                beat["audio_beat_aligned"],
                beat.get("what_happens"),
                beat.get("story_purpose"),
                beat.get("transition_to_next"),
            ]
        )
    char_ws = wb.create_sheet("Characters")
    char_ws.append(["character_id", "source_role_name", "narrative_function", "human_confirmed", "tokens", "beat_presence"])
    for char in project["characters"]:
        char_ws.append(
            [
                char["character_id"],
                char["source_role_name"],
                char["narrative_function"],
                char["human_confirmed"],
                ", ".join(char["identity_description_tokens"]),
                ", ".join(char["beat_presence"]),
            ]
        )
    output = project_dir / "exports" / "analysis.xlsx"
    ensure_dir(output.parent)
    wb.save(output)
    return str(output.resolve())


def main() -> None:
    args = parse_args()
    project_dir = Path(args.project_dir)
    project = load_project(project_dir)
    env = template_env(skill_root())
    analysis_md = env.get_template("analysis_report.j2").render(project=project)
    analysis_md_path = project_dir / "exports" / "analysis.md"
    ensure_dir(analysis_md_path.parent)
    analysis_md_path.write_text(analysis_md, encoding="utf-8")
    analysis_xlsx_path = write_analysis_excel(project_dir, project)
    project["deliverables"]["analysis_markdown_path"] = str(analysis_md_path.resolve())
    project["deliverables"]["analysis_excel_path"] = analysis_xlsx_path
    save_project(project_dir, project)
    print(analysis_md_path.resolve())


if __name__ == "__main__":
    main()
